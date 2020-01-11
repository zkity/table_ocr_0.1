#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import time
import os
from aip import AipOcr
from openpyxl import Workbook


class ocr():
    def __init__(self, app_id, api_key, secret_key, mode, excel_dir):
        self.mode = mode
        self.excel_dir = excel_dir
        self.options = {}
        self.base_dir = '../res/sp'
        self.options["probability"] = "true"
        self.client = AipOcr(app_id, api_key, secret_key)
        if not os.path.exists(excel_dir):
            os.makedirs(excel_dir)

    def __get_file_content__(self, filePath):
        with open(filePath, 'rb') as fp:
            return fp.read()

    def recon(self):
        if not os.path.exists(self.base_dir):
            return
        tables = os.listdir(self.base_dir)
        for table in tables:
            self.__pic_ocr__(table, self.excel_dir)

    def __pic_ocr__(self, table, excel_dir):
        pics_dir = os.path.join(self.base_dir, table)
        pics = os.listdir(pics_dir)
        res = []
        cell_num = len(pics)
        print('recon pic: {}, total {} cells.'.format(table, cell_num))
        i = 0
        for pic in pics:
            i = i + 1
            pic_path = os.path.join(self.base_dir, table, pic)
            image = self.__get_file_content__(pic_path)
            ostr = '\tprocessed {} cells, finish {: .2f}%\r'.format(
                    i, i/cell_num*100)
            print(ostr, end="")
            if self.mode == '1':
                results = self.client.basicAccurate(image, self.options)
            else:
                results = self.client.basicGeneral(image, self.options)
            if 'words_result' in results:
                strx = results.get('words_result')
                if len(strx) > 0:
                    strx = strx[0]
                    v = strx.get('words')
                else:
                    v = 'fail'
            else:
                v = 'fail'

            sp = pic.split('_')
            r = sp[1]
            c = sp[2]
            res.append([r, c, v])
            if self.mode == '1':
                time.sleep(1)
            else:
                time.sleep(0.2)

        # write to file
        wb = Workbook()
        ws = wb.create_sheet('data')

        for row in res:
            r = int(row[0])
            c = int(row[1])
            v = row[2]
            ws.cell(row=r, column=c).value = v

        excel_path = os.path.join(self.excel_dir, '{}.xlsx'.format(table))
        if os.path.exists(excel_path):
            os.remove(excel_path)
        print('\twrite excel file: {}'.format(excel_path))
        wb.save(excel_path)
